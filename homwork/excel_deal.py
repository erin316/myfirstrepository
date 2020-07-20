from openpyxl import load_workbook
from openpyxl import Workbook

#读取数据
workbook = load_workbook(filename="原始成绩.xlsx")

ss=workbook.sheetnames
print(ss)
sheet=workbook[ss[0]]
print(sheet)
#查看表格维度
tt=sheet.dimensions
print(tt)
cell=sheet[tt]
x=[]
for i in cell:
    y=[]
    for j in i:
        y.append(j.value)
    x.append(y)

#清洗数据，将清洗好的数据写入表格
workbook=Workbook()
sheet1=workbook.active
sheet1.title="修改后的表格"
sheet1.append(["学号","姓名","检测","讨论","成绩"])

for item in range(1,len(x)):
    list01=x[item]
    number=list01[0][5:16]
    name=list01[0][16:]
    test=float(list01[4])
    if list01[5]=="-":
        disc=0.0
    else:
        disc=float(list01[5])
    score=float(list01[6])
    final_list=[number,name,test,disc,score]
    sheet1.append(final_list)

workbook.save(filename="作业.xlsx")








